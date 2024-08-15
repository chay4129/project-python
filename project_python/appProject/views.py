import openpyxl
import pandas as pd
from django.shortcuts import render

from .forms import MyImageForm
from django.http import HttpResponse,JsonResponse
from django.core.files.storage import FileSystemStorage
from .models import ExcelFile
from urllib.parse import unquote

from reportlab.pdfgen import canvas
import json
from fpdf import FPDF

def index(request):
    return render(request, 'appProject/pro.html')


def upload_excel(request):
    if request.method == 'POST':
        print("Im inn the post view")
        print(request.POST)
        print(request.FILES)
        form = MyImageForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data['file']
            uploaded_file_instance = ExcelFile.objects.create(file=uploaded_file)
            form.save()

            path = uploaded_file_instance.file.url
            file_path = unquote(path)

            # Open the uploaded Excel file and get the number of sheets
            file = openpyxl.load_workbook(uploaded_file_instance.file.path)
            num_sheets = len(file.sheetnames)


            sheet_names = file.sheetnames
            
            for sheet_name in sheet_names:
                sheet = file[sheet_name]
                result = process_excel_file('uploads/excel/ex.xlsx')
            json = {'file_path': file_path, 'num_sheets': num_sheets,
                    'result': result,
                    'sheet_names': sheet_names}
            js = FPDF.chapter_body(json)
            PDF.create_pdf_from_json(js, "C:/Users/ch/Desktop/new/output.pdf")

            return JsonResponse({'file_path': file_path, 'num_sheets': num_sheets,'result': result, 'sheet_names': sheet_names}, status=200)
        else:
            return HttpResponse("Something went wrong :(  ")
    else:
        print("im in get method")
        return render(request, "appProject/upload_excel.html", {'form': MyImageForm()})




def process_excel_file(file_path):
   try:
       excel_data = pd.ExcelFile(file_path)
       all_sheets_data = []


       for sheet_name in excel_data.sheet_names:
           sheet = excel_data.parse(sheet_name)
           if sheet.empty:
               operation = "Nall"
               columns = []
           else:
               operation = "Sum" if sheet.sum().any().all() else "Average"
               columns = list(sheet.columns)


           sheet_data = {
               "sheet_name": sheet_name,
               "operation": operation,
               "columns": columns
           }
           all_sheets_data.append(sheet_data)


       return all_sheets_data
   except Exception as e:
       return {"error": str(e)}


class PDF(FPDF):
 def create_pdf_from_json(json_data, output_path):
    c = canvas.Canvas(output_path)

    # Convert JSON data to string and write it to the PDF
    json_str = json.dumps(json_data, indent=4)
    c.drawString(100, 800, json_str)

    c.save()


def chapter_body(self, data):
    self.set_font('Arial', '', 12)

    for key, value in data.items():
        self.cell(0, 10, f"{key}: {value}", 0, 1)

    self.ln(10)

#
